"""Buyer excel (Pardavimai) generation for UAB People Link."""
import io
import logging

import httpx
import openpyxl
from openpyxl import Workbook

from app.config import settings
from app.sharepoint import _get_token

logger = logging.getLogger(__name__)

_GRAPH = "https://graph.microsoft.com/v1.0"

MONTH_TITLE = {
    1: "Sausis", 2: "Vasaris", 3: "Kovas", 4: "Balandis",
    5: "Gegužė", 6: "Birželis", 7: "Liepa", 8: "Rugpjūtis",
    9: "Rugsėjis", 10: "Spalis", 11: "Lapkritis", 12: "Gruodis",
}
MONTH_NUM = {v: k for k, v in MONTH_TITLE.items()}

# Pirkėjai_2025.xlsx, "Sąskaitos klientams/Pirkėjų sąskaitos_2025-2026/" —
# addressed by driveId+itemId (not by drive-name/path lookup): the library's
# name doesn't cleanly resolve via SharePoint's diacritic-stripped URL slugs,
# and item IDs are stable across edits to this same file.
SOURCE_DRIVE_ID = "b!MmeQ4_JeWUGthWFPIGG502xnxxzpDkFFlCVzdE5VLn0sFK8Dvk_ZSL1wah8Jo0dY"
SOURCE_ITEM_ID = "01GQTUBMNWYFHH4N4QWJGKPXN7AWVG6IPU"
ARCHIVE_YEAR = 2025
OUTPUT_HEADER = ["Data", "Serija ir numeris", "Suma be PVM", "PVM", "Viso", "Pirkėjo pavadinimas", "Kodas", "PVM kodas"]

FILE_TYPES = {
    "21": "pardavimai 21% PVM",
    "uzsienis": "pardavimai PVM netaikomas (užsienis)",
    "mokymai": "pardavimai PVM netaikomas (mokymai)",
}
FILE_TYPE_LABELS = {
    "21": "21% PVM",
    "uzsienis": "PVM netaikomas (užsienis)",
    "mokymai": "PVM netaikomas (mokymai)",
}


async def download_source_excel() -> bytes:
    token = await _get_token(
        settings.sharepoint_tenant_id,
        settings.sharepoint_client_id,
        settings.sharepoint_client_secret,
    )
    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
        resp = await client.get(
            f"{_GRAPH}/drives/{SOURCE_DRIVE_ID}/items/{SOURCE_ITEM_ID}/content",
            headers={"Authorization": f"Bearer {token}"},
        )
        resp.raise_for_status()
        return resp.content


def load_workbook(content: bytes):
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True)


def parse_month_sheet(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header = [str(c.value or "").strip() for c in ws[1]]

    def col(name: str):
        return header.index(name) if name in header else None

    idx = {
        "data": col("Data"),
        "serija": col("Serija ir numeris"),
        "suma": col("Suma be PVM"),
        "pvm": col("PVM"),
        "viso": col("Viso"),
        "pirkejas": col("Pirkėjo pavadinimas"),
        "kodas": col("Kodas"),
        "pvm_kodas": col("PVM kodas"),
    }
    if idx["serija"] is None:
        return []

    def cell(row, key, default=""):
        i = idx[key]
        return row[i] if i is not None else default

    def text_cell(row, key, default=""):
        val = str(cell(row, key, default) or default).strip()
        return val or default

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        serija = text_cell(row, "serija")
        if not serija:
            continue
        rows.append({
            "data": cell(row, "data", None),
            "serija": serija,
            "suma": cell(row, "suma", 0),
            "pvm": cell(row, "pvm", 0),
            "viso": cell(row, "viso", 0),
            "pirkejas": text_cell(row, "pirkejas"),
            "kodas": text_cell(row, "kodas", "-"),
            "pvm_kodas": text_cell(row, "pvm_kodas", "-"),
        })
    return rows


def get_archive_sheets(wb) -> dict[str, list[tuple]]:
    """Verbatim copy of every '{ARCHIVE_YEAR} [Mėnuo]' sheet, ordered Gruodis→Sausis."""
    prefix = f"{ARCHIVE_YEAR} "
    names = [n for n in wb.sheetnames if n.startswith(prefix)]
    names.sort(key=lambda n: MONTH_NUM.get(n[len(prefix):], 0), reverse=True)

    result = {}
    for name in names:
        ws = wb[name]
        result[name] = list(ws.iter_rows(
            min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True,
        ))
    return result


def classify(row: dict) -> str:
    try:
        pvm = float(row["pvm"] or 0)
    except (TypeError, ValueError):
        pvm = 0
    if pvm > 0.001:
        return "21"
    pvm_kodas = row["pvm_kodas"].strip().upper()
    if not pvm_kodas or pvm_kodas == "-" or pvm_kodas.startswith("LT"):
        return "mokymai"
    return "uzsienis"


def series_of(row: dict) -> str:
    s = row["serija"].upper()
    return s[:4] if len(s) >= 4 else "PL01"


def summarize(month_rows: list[dict]) -> dict[str, int]:
    counts = {"21": 0, "uzsienis": 0, "mokymai": 0}
    for r in month_rows:
        counts[classify(r)] += 1
    return counts


def _write_data_sheet(wb: Workbook, name: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(name)
    ws.append(OUTPUT_HEADER)
    for r in sorted(rows, key=lambda r: (r["data"] is None, r["data"], r["serija"])):
        ws.append([r["data"], r["serija"], r["suma"], r["pvm"], r["viso"], r["pirkejas"], r["kodas"], r["pvm_kodas"]])
    for cell in ws["A"][1:]:
        if cell.value is not None:
            cell.number_format = "yyyy-mm-dd"


def _write_archive_sheet(wb: Workbook, name: str, raw_rows: list[tuple]) -> None:
    ws = wb.create_sheet(name)
    for raw in raw_rows:
        ws.append(list(raw))
    for cell in ws["E"][1:]:
        if cell.value is not None:
            cell.number_format = "yyyy-mm-dd"


def build_workbook(month_rows: list[dict], archive_sheets: dict[str, list[tuple]], file_type: str) -> bytes:
    type_rows = [r for r in month_rows if classify(r) == file_type]

    if file_type == "21":
        series_list = ["PL01", "PL02", "PL03", "PL04"]
    else:
        series_list = sorted({series_of(r) for r in type_rows}) or ["PL01"]

    wb = Workbook()
    wb.remove(wb.active)

    for series in series_list:
        _write_data_sheet(wb, series, [r for r in type_rows if series_of(r) == series])

    for name, raw_rows in archive_sheets.items():
        _write_archive_sheet(wb, name, raw_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def get_month_data(month: int):
    """Returns (month_rows, archive_sheets) for the given 2026 month."""
    content = await download_source_excel()
    wb = load_workbook(content)
    sheet_name = f"2026 {MONTH_TITLE[month]}"
    month_rows = parse_month_sheet(wb, sheet_name)
    archive_sheets = get_archive_sheets(wb)
    return month_rows, archive_sheets


async def generate(month: int, file_type: str) -> tuple[bytes, str]:
    month_rows, archive_sheets = await get_month_data(month)
    xlsx_bytes = build_workbook(month_rows, archive_sheets, file_type)
    filename = f"2026 {month:02d} {FILE_TYPES[file_type]}.xlsx"
    return xlsx_bytes, filename
